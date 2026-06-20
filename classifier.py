import os
from pathlib import Path
from pypdf import PdfReader
from docx import Document
from openpyxl import load_workbook


def read_file_content(filepath):
    try:
        ext = Path(filepath).suffix.lower()

        if ext == ".pdf":
            reader = PdfReader(filepath)
            return "\n".join(page.extract_text() or "" for page in reader.pages[:3])

        if ext == ".docx":
            doc = Document(filepath)
            return "\n".join(p.text for p in doc.paragraphs[:50])

        if ext in [".xlsx", ".xlsm"]:
            wb = load_workbook(filepath, read_only=True, data_only=True)
            text = []
            for sheet in wb.worksheets[:2]:
                for row in sheet.iter_rows(max_row=20, values_only=True):
                    text.append(" ".join(str(c) for c in row if c))
            return "\n".join(text)

        if ext == ".txt":
            with open(filepath, "r", encoding="utf-8", errors="ignore") as f:
                return f.read(3000)

    except Exception as e:
        print(f"[classifier] Could not read content: {e}")

    return ""


def match_keywords(text, keywords):
    text = text.lower()
    return any(k.lower() in text for k in keywords)


def classify_file(filename, config, filepath):
    text = f"{filename}\n{read_file_content(filepath)}".lower()

    base_folder = config["base_folder"]

    for project_name, project_data in config["projects"].items():
        project_keywords = project_data.get("keywords", [])

        if match_keywords(text, project_keywords):
            folders = project_data.get("folders", {})

            for folder_name, folder_keywords in folders.items():
                if folder_keywords and match_keywords(text, folder_keywords):
                    return (
                        project_name,
                        os.path.join(base_folder, project_name, folder_name),
                        f"matched folder {folder_name}"
                    )

            return (
                project_name,
                os.path.join(base_folder, project_name, "General"),
                "matched project only"
            )

    return None, None, "no match"